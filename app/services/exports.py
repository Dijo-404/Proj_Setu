from __future__ import annotations

from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import Image, PageBreak, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import barcode
from barcode.writer import ImageWriter
from PIL import Image as PILImage

from app.models import Batch, InventoryTransaction, ScanLog, Serial
from app.services.log_fields import barcode_sold_by, invoice_created_by, product_audited_by

DEFAULT_LABEL_ROWS = 7
DEFAULT_LABEL_COLUMNS = 3
MIN_LABEL_ROWS = 1
MAX_LABEL_ROWS = 20
MIN_LABEL_COLUMNS = 1
MAX_LABEL_COLUMNS = 8
DANGEROUS_SPREADSHEET_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def spreadsheet_safe(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    if text.startswith(DANGEROUS_SPREADSHEET_PREFIXES):
        return f"'{text}"
    return text


def safe_row(values):
    return [spreadsheet_safe(value) for value in values]


def barcode_png(value: str) -> bytes:
    """Render a Code128 barcode (with the value as human-readable text) as PNG bytes."""
    writer = ImageWriter()
    code = barcode.get("code128", value, writer=writer)
    stream = BytesIO()
    code.write(stream, options={"module_height": 12.0, "font_size": 10, "text_distance": 3.0, "quiet_zone": 2.0})
    return stream.getvalue()


def scans_xlsx(scans: list[ScanLog]) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Scans"
    sheet.append(["Date", "User", "Action", "Serial", "Status", "Batch", "Message", "Tally Reference"])
    for scan in scans:
        sheet.append(
            safe_row(
                [
                    scan.created_at.isoformat(),
                    scan.user.username,
                    scan.action,
                    scan.serial_number_raw,
                    scan.status,
                    scan.batch.batch_number if scan.batch else "",
                    scan.message or "",
                    scan.tally_reference or "",
                ]
            )
        )
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 40)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def serials_xlsx(serials: list[Serial]) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Barcodes"
    sheet.append(
        [
            "Product Code",
            "Product Name",
            "Tally Stock Item",
            "Serial Number",
            "Product Batch",
            "Mfg Date",
            "Expiry Date",
            "Warehouse",
            "Status",
            "Created At",
        ]
    )
    for serial in serials:
        product = serial.product
        sheet.append(
            safe_row(
                [
                    product.product_code,
                    product.product_name,
                    product.tally_stock_item_name,
                    serial.serial_number,
                    serial.product_batch_number or "",
                    serial.mfg_date.isoformat() if serial.mfg_date else "",
                    serial.expiry_date.isoformat() if serial.expiry_date else "",
                    serial.warehouse or "",
                    serial.status,
                    serial.created_at.isoformat(),
                ]
            )
        )
    _autosize(sheet)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def transactions_xlsx(transactions: list[InventoryTransaction]) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"
    sheet.append(
        [
            "Date",
            "User",
            "Type",
            "Serial",
            "Product Code",
            "Product Name",
            "Invoice Created By",
            "Barcode Sold By",
            "Product Audited By",
            "From Status",
            "To Status",
            "Reason",
            "Batch/Reference",
            "Tally Reference",
            "Notes",
        ]
    )
    for txn in transactions:
        sheet.append(
            safe_row(
                [
                    txn.created_at.isoformat(),
                    txn.user.username,
                    txn.transaction_type,
                    txn.serial_number or "",
                    txn.product.product_code if txn.product else "",
                    txn.product.product_name if txn.product else "",
                    invoice_created_by(txn),
                    barcode_sold_by(txn),
                    product_audited_by(txn),
                    txn.status_from or "",
                    txn.status_to or "",
                    txn.reason_code or "",
                    txn.reference_number or "",
                    txn.tally_reference or "",
                    txn.notes or "",
                ]
            )
        )
    _autosize(sheet)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _autosize(sheet) -> None:
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 42)


def label_layout(rows_per_page: int = DEFAULT_LABEL_ROWS, columns_per_page: int = DEFAULT_LABEL_COLUMNS) -> tuple[int, int]:
    rows = rows_per_page if MIN_LABEL_ROWS <= rows_per_page <= MAX_LABEL_ROWS else DEFAULT_LABEL_ROWS
    columns = columns_per_page if MIN_LABEL_COLUMNS <= columns_per_page <= MAX_LABEL_COLUMNS else DEFAULT_LABEL_COLUMNS
    return rows, columns


def _label_image(value: str, target_width: float, target_height: float | None = None) -> Image:
    png = barcode_png(value)
    px_w, px_h = PILImage.open(BytesIO(png)).size
    scale = target_width / px_w
    if target_height is not None:
        scale = min(scale, target_height / px_h)
    width = px_w * scale
    height = px_h * scale
    return Image(BytesIO(png), width=width, height=height)


def barcode_labels_pdf(
    serials: list[Serial],
    rows_per_page: int = DEFAULT_LABEL_ROWS,
    columns_per_page: int = DEFAULT_LABEL_COLUMNS,
) -> bytes:
    rows_per_page, columns_per_page = label_layout(rows_per_page, columns_per_page)
    stream = BytesIO()
    doc = SimpleDocTemplate(stream, pagesize=A4, rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    story = []
    labels_per_page = rows_per_page * columns_per_page
    col_width = doc.width / columns_per_page
    row_height = doc.height / rows_per_page
    image_width = max(col_width - 4 * mm, col_width * 0.7)
    image_height = max(row_height - 4 * mm, row_height * 0.7)

    for page_start in range(0, len(serials), labels_per_page):
        page_serials = serials[page_start:page_start + labels_per_page]
        table_rows = []
        for row_index in range(rows_per_page):
            row = []
            for column_index in range(columns_per_page):
                serial_index = row_index * columns_per_page + column_index
                if serial_index < len(page_serials):
                    row.append(_label_image(page_serials[serial_index].serial_number, image_width, image_height))
                else:
                    row.append("")
            table_rows.append(row)
        table = Table(
            table_rows,
            colWidths=[col_width] * columns_per_page,
            rowHeights=[row_height] * rows_per_page,
        )
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 2 * mm),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2 * mm),
                    ("TOPPADDING", (0, 0), (-1, -1), 2 * mm),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2 * mm),
                ]
            )
        )
        story.append(table)
        if page_start + labels_per_page < len(serials):
            story.append(PageBreak())
    if not story:
        styles = getSampleStyleSheet()
        story = [Paragraph("No labels selected", styles["BodyText"])]
    doc.build(story)
    return stream.getvalue()


def audit_report_pdf(batch: Batch) -> bytes:
    stream = BytesIO()
    doc = SimpleDocTemplate(stream, pagesize=A4, rightMargin=14 * mm, leftMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"Audit Report: {batch.batch_number}", styles["Title"]),
        Paragraph(f"Reference: {batch.party_name or '-'}", styles["BodyText"]),
        Paragraph(f"Status: {batch.status}", styles["BodyText"]),
        Spacer(1, 5 * mm),
    ]
    rows = [["Finding", "Serial", "Product", "Expected", "Scanned"]]
    for finding in batch.audit_findings:
        rows.append(
            [
                finding.finding_type,
                finding.serial_number,
                finding.product_name or "",
                finding.expected_status or "",
                finding.scanned_status or "",
            ]
        )
    table = Table(rows, repeatRows=1, colWidths=[28 * mm, 42 * mm, 58 * mm, 28 * mm, 28 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8f2ff")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    return stream.getvalue()
