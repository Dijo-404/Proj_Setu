from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Batch, BatchItem, BatchStatus, BatchType, Product, ScanLog, Serial, SerialStatus, TransactionType, User, WarehouseLevel
from app.services.expiry import parse_optional_date
from app.services.inventory import InventoryError, create_batch, generate_serials, log_inventory_transaction, normalize_serial


MAX_ASSIGNMENT_QUANTITY = 5000


@dataclass(frozen=True)
class AssignmentLine:
    product: Product
    quantity: int
    prefix: str | None = None
    product_batch_number: str | None = None
    mfg_date: date | None = None
    expiry_date: date | None = None
    warehouse: str | None = None
    warehouse_level: str = WarehouseLevel.COMPANY_WAREHOUSE.value


def parse_bulk_assignment_xlsx(db: Session, data: bytes) -> list[AssignmentLine]:
    try:
        workbook = load_workbook(BytesIO(data), data_only=True)
    except Exception as exc:
        raise InventoryError("Upload a readable Excel .xlsx file") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise InventoryError("Excel file is empty")

    header = [str(value or "").strip().lower().replace("_", " ") for value in rows[0]]
    product_col = _find_column(header, {"product code", "code", "product"})
    qty_col = _find_column(header, {"quantity", "qty"})
    batch_col = _find_column(header, {"batch", "batch no", "batch number", "product batch"})
    mfg_col = _find_column(header, {"mfg date", "manufacturing date", "manufacture date"})
    expiry_col = _find_column(header, {"expiry date", "expiry", "exp date"})
    warehouse_col = _find_column(header, {"warehouse", "wh", "location"})
    warehouse_level_col = _find_column(header, {"warehouse level", "franchise level", "location level"})
    if product_col is None or qty_col is None:
        product_col, qty_col = 0, 1

    lines: list[AssignmentLine] = []
    total = 0
    for index, row in enumerate(rows[1:], start=2):
        raw_code = row[product_col] if product_col < len(row) else None
        raw_qty = row[qty_col] if qty_col < len(row) else None
        code = normalize_serial(str(raw_code or ""))
        if not code and raw_qty in {None, ""}:
            continue
        if not code:
            raise InventoryError(f"Row {index}: product code is required")
        try:
            quantity = int(raw_qty)
        except (TypeError, ValueError) as exc:
            raise InventoryError(f"Row {index}: quantity must be a whole number") from exc
        if quantity < 1:
            raise InventoryError(f"Row {index}: quantity must be at least 1")
        product = db.scalar(select(Product).where(Product.product_code == code, Product.active == True))
        if not product:
            raise InventoryError(f"Row {index}: product {code} was not found")
        total += quantity
        try:
            mfg_date = parse_optional_date(row[mfg_col]) if mfg_col is not None and mfg_col < len(row) else None
            expiry_date = parse_optional_date(row[expiry_col]) if expiry_col is not None and expiry_col < len(row) else None
        except (TypeError, ValueError) as exc:
            raise InventoryError(f"Row {index}: use a valid date for mfg/expiry") from exc
        if mfg_date and expiry_date and expiry_date <= mfg_date:
            raise InventoryError(f"Row {index}: expiry date must be after mfg date")
        product_batch_number = str(row[batch_col] or "").strip() if batch_col is not None and batch_col < len(row) else None
        warehouse = str(row[warehouse_col] or "").strip() if warehouse_col is not None and warehouse_col < len(row) else None
        warehouse_level = (
            str(row[warehouse_level_col] or "").strip()
            if warehouse_level_col is not None and warehouse_level_col < len(row)
            else WarehouseLevel.COMPANY_WAREHOUSE.value
        )
        try:
            warehouse_level = WarehouseLevel(warehouse_level or WarehouseLevel.COMPANY_WAREHOUSE.value).value
        except ValueError as exc:
            raise InventoryError(f"Row {index}: warehouse level is not recognized") from exc
        lines.append(
            AssignmentLine(
                product=product,
                quantity=quantity,
                product_batch_number=product_batch_number or None,
                mfg_date=mfg_date,
                expiry_date=expiry_date,
                warehouse=warehouse or None,
                warehouse_level=warehouse_level,
            )
        )

    if not lines:
        raise InventoryError("Excel file has no assignment rows")
    if total > MAX_ASSIGNMENT_QUANTITY:
        raise InventoryError(f"Assign {MAX_ASSIGNMENT_QUANTITY} barcodes or fewer at a time")
    return lines


def assign_barcodes_to_existing_stock(
    db: Session,
    user: User,
    lines: list[AssignmentLine],
    notes: str | None = None,
    source: str = "MANUAL",
    initial_status: SerialStatus = SerialStatus.IN_STOCK,
) -> Batch:
    total = sum(line.quantity for line in lines)
    if total < 1:
        raise InventoryError("Quantity must be at least 1")
    if total > MAX_ASSIGNMENT_QUANTITY:
        raise InventoryError(f"Assign {MAX_ASSIGNMENT_QUANTITY} barcodes or fewer at a time")

    try:
        batch = create_batch(
            db,
            user,
            BatchType.QR_ASSIGNMENT,
            party_name="Existing Tally stock",
            notes=notes,
            reason_code=source,
            commit=False,
        )
        created_serials: list[Serial] = []
        for line in lines:
            created_serials.extend(
                generate_serials(
                    db,
                    line.product,
                    line.quantity,
                    prefix=line.prefix or line.product.product_code,
                    initial_status=initial_status,
                    product_batch_number=line.product_batch_number,
                    mfg_date=line.mfg_date,
                    expiry_date=line.expiry_date,
                    warehouse=line.warehouse,
                    warehouse_level=line.warehouse_level,
                    commit=False,
                )
            )

        message = (
            "Barcode assigned to existing Tally stock"
            if initial_status == SerialStatus.IN_STOCK
            else "Barcode generated for future stock movement"
        )
        for serial in created_serials:
            db.add(BatchItem(batch_id=batch.id, serial_id=serial.id))
            db.add(
                ScanLog(
                    serial_id=serial.id,
                    serial_number_raw=serial.serial_number,
                    user_id=user.id,
                    action=TransactionType.QR_ASSIGNMENT.value,
                    batch_id=batch.id,
                    status=initial_status.value,
                    message=message,
                )
            )
            log_inventory_transaction(
                db,
                user,
                TransactionType.QR_ASSIGNMENT,
                serial=serial,
                product=serial.product,
                batch=batch,
                status_from=None,
                status_to=initial_status.value,
                reason_code=source,
                notes=notes or message,
            )
        batch.status = BatchStatus.CLOSED.value
        batch.submitted_at = batch.created_at
        batch.synced_at = batch.created_at
        db.commit()
    except Exception:
        db.rollback()
        raise
    db.refresh(batch)
    return batch


def _find_column(header: list[str], names: set[str]) -> int | None:
    for index, value in enumerate(header):
        if value in names:
            return index
    return None
