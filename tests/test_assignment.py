from io import BytesIO
from types import SimpleNamespace

from openpyxl import Workbook
from sqlalchemy import func, select

from app.models import Batch, BatchItem, BatchStatus, BatchType, ChangeAudit, InventoryTransaction, Product, Serial, SerialStatus, TransactionType, User
from app.services import assignment as assignment_service
from app.services.assignment import AssignmentLine, assign_barcodes_to_existing_stock, parse_bulk_assignment_xlsx
from app.services.exports import serials_xlsx
from app.templates import templates


def make_product(code="SG080"):
    return Product(
        product_code=code,
        product_name="Turmeric Powder 100g",
        hsn="0910",
        gst_rate=5,
        unit="Pcs",
        default_rate=100,
        tally_stock_item_name="SG Turmeric Powder 100g",
    )


def test_assign_barcodes_to_existing_stock_creates_history(db_session):
    user = User(username="admin", password_hash="x", role="admin")
    product = make_product()
    db_session.add_all([user, product])
    db_session.commit()

    batch = assign_barcodes_to_existing_stock(db_session, user, [AssignmentLine(product=product, quantity=3)])

    assert batch.batch_type == BatchType.QR_ASSIGNMENT.value
    assert batch.status == BatchStatus.CLOSED.value
    assert len(batch.items) == 3
    assert {item.serial.status for item in batch.items} == {SerialStatus.IN_STOCK.value}
    transactions = db_session.scalars(select(InventoryTransaction).order_by(InventoryTransaction.id)).all()
    assert [txn.transaction_type for txn in transactions] == [TransactionType.QR_ASSIGNMENT.value] * 3
    assert {txn.status_to for txn in transactions} == {SerialStatus.IN_STOCK.value}


def test_bulk_assignment_xlsx_uses_existing_product_codes(db_session):
    product = make_product("SG081")
    db_session.add(product)
    db_session.commit()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Product Code", "Quantity"])
    sheet.append(["SG081", 2])
    stream = BytesIO()
    workbook.save(stream)

    lines = parse_bulk_assignment_xlsx(db_session, stream.getvalue())

    assert len(lines) == 1
    assert lines[0].product.id == product.id
    assert lines[0].quantity == 2


def test_bulk_assignment_xlsx_can_create_product_from_imported_name_and_tax(db_session):
    user = User(username="import-admin", password_hash="x", role="admin")
    db_session.add(user)
    db_session.commit()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Product Name", "HSN", "GST", "SGST", "IGST", "Quantity"])
    sheet.append(["SG Omega Softgel 500mg", "2106", "5", "2.5", "", 3])
    stream = BytesIO()
    workbook.save(stream)

    lines = parse_bulk_assignment_xlsx(db_session, stream.getvalue(), user=user)

    assert len(lines) == 1
    assert lines[0].quantity == 3
    assert lines[0].product.product_name == "SG Omega Softgel 500mg"
    assert lines[0].product.hsn == "2106"
    assert lines[0].product.gst_rate == 5
    assert lines[0].product.tally_stock_item_name == "SG Omega Softgel 500mg"
    audit = db_session.scalar(select(ChangeAudit).where(ChangeAudit.entity_type == "product"))
    assert audit is not None
    assert audit.actor_username == "import-admin"


def test_bulk_assignment_xlsx_matches_product_alias_names(db_session):
    product = make_product("SGALIAS")
    product.nickname = "Customer Friendly Alias"
    product.alternate_tally_stock_item_name = "Tally Alias Two"
    db_session.add(product)
    db_session.commit()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Product Name", "Quantity"])
    sheet.append(["Tally Alias Two", 2])
    stream = BytesIO()
    workbook.save(stream)

    lines = parse_bulk_assignment_xlsx(db_session, stream.getvalue())

    assert len(lines) == 1
    assert lines[0].product.id == product.id
    assert lines[0].quantity == 2


def test_assignment_page_has_searchable_alias_product_selector():
    product = SimpleNamespace(
        id=1,
        product_code="SGALIAS",
        product_name="Formal Product",
        nickname="Friendly Alias",
        tally_stock_item_name="Primary Tally",
        alternate_tally_stock_item_name="Second Tally",
    )

    html = templates.env.get_template("barcode_assignment.html").render(
        user=None,
        products=[product],
        batches=[],
        warehouse_levels=["Company Warehouse"],
        error=None,
    )

    assert 'id="assignment-product-search"' in html
    assert 'id="assignment-product-select"' in html
    assert "Friendly Alias" in html
    assert "Second Tally" in html


def test_tally_invoice_export_import_generates_assignment_labels(db_session):
    user = User(username="tally-import-admin", password_hash="x", role="admin")
    db_session.add(user)
    db_session.commit()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "INVOICE"

    def append_row(values):
        row = [None] * 18
        for col, value in values.items():
            row[col - 1] = value
        sheet.append(row)

    append_row({6: "INVOICE"})
    append_row({1: "SWARNAGOWRI 26-27", 10: "Invoice No.", 13: "Dated"})
    append_row({1: "Supplier (Bill from)"})
    append_row({1: "KAI HERBALS"})
    append_row({1: "Sl", 2: "Description of Goods", 11: "Quantity", 12: "Rate", 13: "per", 15: "Amount"})
    append_row({})
    append_row({1: 1, 2: "SG OMEGA 3SOFT GEL 500 MG(60 CAP)", 11: 3, 12: 95.24, 13: "Pac", 15: 285.72})
    append_row({1: 2, 2: "SG OMEGA SOFTGELCAPSULE 1000MG(30 CAP)", 11: 4, 12: 114.29, 13: "Pac", 15: 457.16})
    append_row({15: 742.88})
    append_row({3: "OUTPUT IGST @ 5 %", 12: 5, 13: "%", 15: 37.14})
    append_row({3: "ROUND OFF", 15: -0.02})
    append_row({2: "Total", 11: 7, 15: 780})
    stream = BytesIO()
    workbook.save(stream)

    lines = parse_bulk_assignment_xlsx(db_session, stream.getvalue(), user=user)
    batch = assign_barcodes_to_existing_stock(db_session, user, lines, source="BULK_EXCEL")

    assert [line.quantity for line in lines] == [3, 4]
    assert {line.product.gst_rate for line in lines} == {5}
    assert {line.product.unit for line in lines} == {"Pac"}
    assert len(batch.items) == 7
    assert db_session.scalar(select(func.count(Product.id))) == 2
    assert db_session.scalar(select(func.count(Serial.id))) == 7
    assert db_session.scalar(select(func.count(BatchItem.id))) == 7


def test_serials_xlsx_exports_generated_barcodes(db_session):
    user = User(username="admin", password_hash="x", role="admin")
    product = make_product("SG082")
    db_session.add_all([user, product])
    db_session.commit()
    batch = assign_barcodes_to_existing_stock(db_session, user, [AssignmentLine(product=product, quantity=1)])

    data = serials_xlsx([batch.items[0].serial])

    assert data.startswith(b"PK")


def test_assignment_rolls_back_every_record_when_history_write_fails(db_session, monkeypatch):
    user = User(username="atomic-admin", password_hash="x", role="admin")
    first = make_product("SG083")
    second = make_product("SG084")
    db_session.add_all([user, first, second])
    db_session.commit()

    def fail_history(*_args, **_kwargs):
        raise RuntimeError("simulated history failure")

    monkeypatch.setattr(assignment_service, "log_inventory_transaction", fail_history)

    try:
        assign_barcodes_to_existing_stock(
            db_session,
            user,
            [AssignmentLine(product=first, quantity=1), AssignmentLine(product=second, quantity=1)],
        )
    except RuntimeError:
        pass
    else:
        assert False, "the simulated failure must escape"

    assert db_session.scalar(select(func.count(Batch.id))) == 0
    assert db_session.scalar(select(func.count(Serial.id))) == 0
    assert db_session.scalar(select(func.count(BatchItem.id))) == 0
    assert db_session.scalar(select(func.count(InventoryTransaction.id))) == 0
