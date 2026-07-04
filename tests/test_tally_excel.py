from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.models import BatchItem, BatchType, Product, ScanLog, SerialStatus, User
from app.services.assignment import parse_bulk_assignment_xlsx
from app.services.inventory import InventoryError, add_serial_to_batch, create_batch, generate_serials
from app.services.sale_returns import scan_sale_return_product
from app.services.tally_excel import TALLY_ACCOUNTING_VOUCHER_HEADERS, batch_tally_xlsx, import_tally_excel_to_batch


VALID_TALLY_EXCEL_SETTINGS = {
    "sales_voucher_type": "Sales",
    "purchase_voucher_type": "Purchase",
    "sales_ledger_name": "Sales Ledger",
    "purchase_ledger_name": "Purchase Ledger",
    "cgst_ledger_name": "Input CGST @  2.5 %",
    "sgst_ledger_name": "Input SGST@2.5%",
    "sales_gst_ledger_mappings": "5 | Sales @ 5% | Output CGST @ 2.5% | Output SGST @ 2.5% | Output IGST @ 5%",
    "round_off_ledger_name": "ROUND OFF",
}


def _workbook_bytes(headers: list[str], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _product(code: str = "TALLYXL") -> Product:
    return Product(
        product_code=code,
        product_name="Tally Excel Product",
        hsn="0910",
        gst_rate=5,
        unit="Pcs",
        default_rate=100,
        tally_stock_item_name="Tally Excel Product",
    )


def test_tally_excel_import_picks_fefo_stock_and_keeps_rate(db_session):
    user = User(username="admin", password_hash="x", role="admin")
    product = _product()
    db_session.add_all([user, product])
    db_session.commit()
    late = generate_serials(
        db_session,
        product,
        1,
        initial_status=SerialStatus.IN_STOCK,
        expiry_date=date(2026, 12, 31),
    )[0]
    early = generate_serials(
        db_session,
        product,
        1,
        initial_status=SerialStatus.IN_STOCK,
        expiry_date=date(2026, 8, 1),
    )[0]
    next_early = generate_serials(
        db_session,
        product,
        1,
        initial_status=SerialStatus.IN_STOCK,
        expiry_date=date(2026, 9, 1),
    )[0]
    batch = create_batch(db_session, user, BatchType.ISSUE, "Marketing", "", "SAMPLE")
    data = _workbook_bytes(
        ["Description of Goods", "Quantity", "Rate"],
        [[product.tally_stock_item_name, 2, 123.45]],
    )

    result = import_tally_excel_to_batch(db_session, batch, user, data)

    items = db_session.scalars(select(BatchItem).where(BatchItem.batch_id == batch.id)).all()
    logs = db_session.scalars(select(ScanLog).where(ScanLog.batch_id == batch.id)).all()
    assert result.quantity == 2
    assert {item.serial_id for item in items} == {early.id, next_early.id}
    assert late.id not in {item.serial_id for item in items}
    assert {item.rate for item in items} == {123.45}
    assert all(item.fefo_picked for item in items)
    assert {log.status for log in logs} == {"EXCEL_IMPORTED"}


def test_tally_excel_export_can_be_read_back_as_import_lines(db_session):
    user = User(username="admin", password_hash="x", role="admin")
    product = _product("TALLYX2")
    db_session.add_all([user, product])
    db_session.commit()
    generate_serials(db_session, product, 2, initial_status=SerialStatus.IN_STOCK)
    batch = create_batch(db_session, user, BatchType.ISSUE, "Marketing", "", "SAMPLE")
    import_tally_excel_to_batch(
        db_session,
        batch,
        user,
        _workbook_bytes(["Product Code", "Quantity", "Rate"], [[product.product_code, 2, 88.5]]),
    )

    data = batch_tally_xlsx(batch)
    workbook = load_workbook(BytesIO(data), data_only=True)
    sheet = workbook.active
    lines = parse_bulk_assignment_xlsx(db_session, data, allow_product_create=False)

    assert data.startswith(b"PK")
    assert sheet["B6"].value == "Description of Goods"
    assert sheet["B7"].value == product.tally_stock_item_name
    assert sheet["F7"].value == 2
    assert sheet["H7"].value == 88.5
    assert len(lines) == 1
    assert lines[0].product.id == product.id
    assert lines[0].quantity == 2
    assert lines[0].rate == Decimal("88.5")


def test_sale_tally_excel_export_uses_tally_accounting_voucher_template(db_session):
    user = User(username="sales-xlsx", password_hash="x", role="sales")
    product = _product("TALLYSALE")
    db_session.add_all([user, product])
    db_session.commit()
    serials = generate_serials(db_session, product, 2, initial_status=SerialStatus.IN_STOCK)
    batch = create_batch(db_session, user, BatchType.SALE, "Customer Ledger", "")
    for serial in serials:
        add_serial_to_batch(db_session, batch, user, serial.serial_number)

    data = batch_tally_xlsx(batch, VALID_TALLY_EXCEL_SETTINGS)
    workbook = load_workbook(BytesIO(data), data_only=True)
    sheet = workbook.active
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    ledger_col = headers.index("Ledger Name") + 1
    amount_col = headers.index("Ledger Amount") + 1
    drcr_col = headers.index("Ledger Amount Dr/Cr") + 1
    item_col = headers.index("Item Name") + 1
    qty_col = headers.index("Billed Quantity") + 1
    rate_col = headers.index("Item Rate") + 1
    hsn_col = headers.index("HSN/SAC") + 1
    change_mode_col = headers.index("Change Mode ") + 1

    assert sheet.title == "Accounting Voucher"
    assert headers == TALLY_ACCOUNTING_VOUCHER_HEADERS
    assert "Description of Goods" not in headers
    assert "HSN Code" not in headers
    assert "GST Rate %" not in headers
    assert sheet.cell(2, ledger_col).value == "Customer Ledger"
    assert sheet.cell(2, amount_col).value == 210
    assert sheet.cell(2, drcr_col).value == "Dr"
    assert sheet.cell(3, ledger_col).value == "Sales @ 5%"
    assert sheet.cell(3, drcr_col).value == "Cr"
    assert sheet.cell(3, item_col).value == product.tally_stock_item_name
    assert sheet.cell(3, qty_col).value == 2
    assert sheet.cell(3, rate_col).value == 100
    assert sheet.cell(3, hsn_col).value == product.hsn
    assert sheet.cell(3, change_mode_col).value == "Accounting Invoice"
    assert sheet.cell(4, ledger_col).value == "Output CGST @ 2.5%"
    assert sheet.cell(5, ledger_col).value == "Output SGST @ 2.5%"


def test_sale_tally_excel_export_has_one_item_row_and_can_be_imported_by_app(db_session):
    user = User(username="sales-xlsx-roundtrip", password_hash="x", role="sales")
    product = _product("TALLYSALE2")
    db_session.add_all([user, product])
    db_session.commit()
    serials = generate_serials(db_session, product, 2, initial_status=SerialStatus.IN_STOCK)
    batch = create_batch(db_session, user, BatchType.SALE, "Customer Ledger", "")
    for serial in serials:
        add_serial_to_batch(db_session, batch, user, serial.serial_number)

    data = batch_tally_xlsx(batch, VALID_TALLY_EXCEL_SETTINGS)
    workbook = load_workbook(BytesIO(data), data_only=True)
    sheet = workbook.active
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    item_col = headers.index("Item Name") + 1
    qty_col = headers.index("Billed Quantity") + 1
    item_rows = [
        row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, item_col).value == product.tally_stock_item_name
    ]
    lines = parse_bulk_assignment_xlsx(db_session, data, allow_product_create=False)

    assert item_rows == [3]
    assert sheet.cell(item_rows[0], qty_col).value == 2
    assert len(lines) == 1
    assert lines[0].product.id == product.id
    assert lines[0].quantity == 2
    assert lines[0].rate == Decimal("100")


def test_tally_excel_import_rejects_purchase_batches(db_session):
    user = User(username="admin", password_hash="x", role="admin")
    product = _product("TALLYX3")
    db_session.add_all([user, product])
    db_session.commit()
    batch = create_batch(db_session, user, BatchType.PURCHASE, "Supplier", "")
    data = _workbook_bytes(["Product Code", "Quantity"], [[product.product_code, 1]])

    with pytest.raises(InventoryError, match="sale, issue, and purchase return"):
        import_tally_excel_to_batch(db_session, batch, user, data)


def test_sale_tally_excel_import_waits_for_pending_return_shelf_scan(db_session):
    user = User(username="sales", password_hash="x", role="sales")
    product = _product("TALLYX4")
    db_session.add_all([user, product])
    db_session.commit()
    serials = generate_serials(db_session, product, 2, initial_status=SerialStatus.IN_STOCK)
    batch = create_batch(db_session, user, BatchType.SALE, "Customer Ledger", "")
    add_serial_to_batch(db_session, batch, user, serials[0].serial_number)
    scan_sale_return_product(db_session, batch, user, serials[0].serial_number)
    data = _workbook_bytes(["Product Code", "Quantity"], [[product.product_code, 1]])

    with pytest.raises(InventoryError, match="shelf QR"):
        import_tally_excel_to_batch(db_session, batch, user, data)

    items = db_session.scalars(select(BatchItem).where(BatchItem.batch_id == batch.id)).all()
    assert items == []
